#!/usr/bin/env python3
"""
new_job_definition.py — 新岗位定义生成

从已标注JD中筛选新岗位候选（是否新岗位候选=='是'），
按归一化岗位名分组，每组聚合成一份结构化岗位定义：
  岗位名称、别称、核心职责、必备技能、加分技能、典型行业应用场景

规则引擎（默认）: 正则+统计，零LLM，确定性输出
LLM兜底（--use_llm）: 规则初稿→LLM精炼核心职责+行业场景→幻觉校验→择优采纳

Usage:
    # 纯规则（默认，零依赖）
    python new_job_definition.py \
        --input_dir "./【已标注】filtered/jd_v2/" \
        --output new_job_definitions.json

    # LLM精炼（需 DEEPSEEK_API_KEY 环境变量）
    python new_job_definition.py \
        --input_dir "./【已标注】filtered/jd_v2/" \
        --output new_job_definitions.json \
        --use_llm --llm_max 30
"""

import os, sys, json, argparse, csv, re, glob, time, ssl
from collections import defaultdict, Counter
from datetime import datetime
from difflib import SequenceMatcher

# SSL context for corporate proxy environments
_SSL_CONTEXT = ssl.create_default_context()
_SSL_CONTEXT.check_hostname = False
_SSL_CONTEXT.verify_mode = ssl.CERT_NONE

# LLM API 配置（与 synthesize_data.py 保持一致）
LLM_CONFIG = {
    'api_base': 'https://api.deepseek.com/v1',
    'model': 'deepseek-chat',
    'fallback_model': 'deepseek-chat',
    'max_tokens': 2048,
    'api_key': 'sk-f2e63f9bbe8d41f4aad7aaa3d8145f8c',
}


# ============================================================
#  常量
# ============================================================

# 可从岗位名中剥离的通用修饰词
STRIP_TOKENS = r'高级|资深|初级|助理|实习|应届|校招|管培|培训生|见习|总监|主管|负责人'

PAREN_SUFFIX = re.compile(r'[（(][^)）]*[)）]$')

# 行业关键词 → 行业名
INDUSTRY_MAP = [
    # ── 从公司名+JD原文匹配行业，优先级从高到低 ──
    (r'汽车|整车|新能源车|自动驾驶|adas|车联网|智能驾驶|无人驾驶|智驾|座舱|车规', '智能汽车/自动驾驶'),
    (r'金融|银行|保险|证券|风控|支付|信用|投融资|基金|信托|期货|理财|信贷|消费金融|fintech', '金融科技'),
    (r'医疗|医药|医院|健康|影像|诊断|基因|临床|制药|药企|患者|病历|诊疗|医疗器械|生物医药|体检|康复|护理|卫健委|疾控', '智慧医疗'),
    (r'制造|工业|工厂|质检|产线|智能制造|精密加工|数控|PLM|MES|工业互联网|工控|工业机器人|工业视觉|产线自动化|工业制造|装备制造|机械制造', '智能制造'),
    (r'教育|培训|课程|教学|院校|在线教育|慕课|MOOC|题库|智慧课堂|校园|教务|自适应学习|学情分析|教育科技|AI助教|智能题库|智慧学习', '智慧教育'),
    (r'电商|零售|物流|仓储|供应链|配送|快递|外卖|O2O|新零售|跨境电商|Shopify|淘宝|天猫|京东|拼多多|美团|饿了么', '电商与新零售'),
    (r'电信|5G|运营商|基站|光通信|核心网|接入网|传输网|联通|广电|中国电信|中国移动|移动通信|通信网络|无线通信|网络通信', '通信网络'),
    (r'游戏|电竞|娱乐|短视频|直播|视频|影视|动漫|网文|小说|出版|传媒|广告|营销|自媒体|AIGC|内容推荐|内容分发|视频理解|AIGC内容|数字人|虚拟人|内容创作|数字娱乐|游戏引擎|游戏开发|游戏策划|抽卡|卡牌|角色设计|原画', '数字内容与娱乐'),
    (r'安防|巡检|应急|消防|公安|警务|军工|国防|保密|密码|等保|护网|攻防|渗透|网络安全|信息安全|数据安全|公共安全', '公共安全与安防'),
    (r'芯片|半导体|IC|集成电路|晶圆|封装|流片|EDA|IP核|FPGA|SOC|ASIC|硅光', '半导体与芯片'),
    (r'能源|电力|电网|光伏|风电|储能|新能源|碳中和|碳达峰|氢能|核能|充电桩|换电|综合能源|智慧能源|电力交易', '能源科技'),
    (r'政务|政府|智慧城市|数字政府|城市大脑|公共|政务云|一网通办|数字孪生城市|社会治理|网格化|城管', '数字政务与智慧城市'),
    (r'招聘|HR|人事|猎头|人力|eHR|HCM|薪酬|绩效|考勤|灵活用工|猎聘|人力资源管理|HRM|HRTech', 'HR科技'),
    (r'地产|建筑|施工|土木工程|建筑工程|工程勘察|工程设计|工程施工|测绘|BIM|智慧工地|物业管理|装修|建材|装配式|绿色建筑|房地产', '建筑与地产科技'),
    (r'农业|种植|养殖|畜牧|渔业|农田|智慧农业|精准农业|农产品|农机|灌溉|大棚|水产', '智慧农业'),
    (r'法律|律师|法务|合规|知产|知识产权|专利|商标|版权|司法|法院|检察院|仲裁', '法律科技'),
    # ── 新增高频行业 ──
    (r'航空|航天|卫星|火箭|无人机|飞行|导航|遥感|北斗|GNS', '航空航天'),
    (r'环保|环境|水务|污水|垃圾|固废|大气|监测|碳中和|碳交易|生态|绿化|环卫', '环保与碳中和'),
    (r'酒店|旅游|民宿|景区|OTA|机票|旅行社|出行|网约车|共享单车|租车', '旅游与出行'),
    (r'餐饮|食品|外卖|生鲜|食材|预制菜|调味品|饮料|乳业|酒类|烘焙', '餐饮与食品科技'),
    (r'宠物|动物|兽医|宠物医院|宠物食品|宠物用品', '宠物经济'),
    (r'运动|体育|健身|瑜伽|跑步|球类|体育赛事|电竞', '体育科技'),
    # ── 公司后缀宽松匹配 ──
    (r'科技|信息|数据|软件|网络|互联|智能|数字|云端|云计|IT|SaaS|PaaS|平台', '通用信息技术'),
    (r'咨询|外包|顾问|企业服务|管理咨询|IT服务|技术服务|专业服务|BPO|SaaS平台|企业级', '企业服务/咨询'),
]

# 行业兜底关键词（JD原文扫描，不依赖公司名）
INDUSTRY_JD_KEYWORDS = {
    '金融科技': ['风控模型', '信用评分', '反欺诈', '量化交易', '智能投顾', '保险科技', '支付清算'],
    '智慧医疗': ['医学影像', '病理诊断', '药物发现', 'AI制药', '电子病历', '远程医疗'],
    '智能制造': ['工业视觉', '缺陷检测', '预测性维护', '数字孪生工厂', '产线自动化'],
    '智能汽车/自动驾驶': ['感知算法', '路径规划', '激光雷达', '毫米波雷达', 'BEV感知', '端到端自动驾驶'],
    '电商与新零售': ['推荐系统', '搜索算法', '价格预测', '库存优化', '智能客服'],
    '数字内容与娱乐': ['内容推荐', '视频理解', 'AIGC内容', '虚拟人', '数字人直播'],
    '通信网络': ['网络优化', '信号处理', '频谱管理', '波束赋形'],
    '半导体与芯片': ['芯片设计', 'EDA工具', '版图设计', '数字IC', '模拟IC', 'RTL设计'],
    '能源科技': ['光伏逆变', '电池管理', '智能电网', '负荷预测', '虚拟电厂'],
    '环保与碳中和': ['碳排放', '碳足迹', '环境监测', '大气模型', '水质预测'],
    '航空航天': ['卫星图像', '遥感解译', '轨道计算', '空间态势感知'],
    '法律科技': ['合同审查', '案例检索', '法律文书', '智能审判'],
    '教育': ['自适应学习', '智能题库', 'AI助教', '学情分析'],
}

# AI信号（用于职责排序 + 子分组）
AI_SIGNALS = [
    '大模型', 'LLM', '大语言模型', 'GPT', 'ChatGPT', 'DeepSeek',
    'AIGC', '生成式AI', 'GenAI',
    'Agent', 'AI Agent', '智能体', 'Multi-Agent', '多智能体',
    'RAG', '检索增强生成', 'Graph RAG',
    '大模型微调', 'SFT', 'RLHF', 'DPO',
    '多模态大模型', '视觉大模型', 'VLM', '多模态',
    'LangChain', 'LangGraph', 'LlamaIndex', 'Dify', 'Coze',
    'Stable Diffusion', 'SDXL', 'Midjourney', 'Sora',
    'LoRA', 'QLoRA', 'PEFT',
    'Prompt Engineering', '提示工程',
    'AI Coding', 'Copilot', 'CodeAgent',
    '数字孪生', '数字人', '具身智能',
    '联邦学习', '隐私计算',
]

# AI信号家族（用于子分组）
AI_FAMILIES = {
    '大模型': ['大模型', 'LLM', '大语言模型', 'GPT', 'ChatGPT', '大模型微调', 'SFT', 'RLHF', 'DPO',
              'DeepSeek', '预训练', '模型训练', '大语言', 'LLaMA', 'Qwen'],
    'Agent': ['Agent', 'AI Agent', '智能体', 'Multi-Agent', '多智能体', 'LangChain', 'LangGraph',
              'LlamaIndex', 'Dify', 'Coze', 'Agentic', 'AI Coding', 'Copilot', 'CodeAgent'],
    'AIGC': ['AIGC', '生成式AI', 'GenAI', 'Stable Diffusion', 'SDXL', 'Midjourney', 'Sora',
             'LoRA', 'QLoRA', '图像生成', '视频生成', '文生图', '文生视频'],
    '多模态': ['多模态大模型', '视觉大模型', 'VLM', '多模态', '视觉语言'],
    'RAG': ['RAG', '检索增强生成', 'Graph RAG', '向量数据库', 'Embedding'],
}


# ============================================================
#  岗位名归一化
# ============================================================

# ── 段打分组件（normalize_job_name 使用的辅助函数） ──

# 正向信号：有这些词 → 像是岗位名
_JOB_SUFFIX_RE = re.compile(r'(工程师|研究员|分析师|设计师|架构师|经理|专家|总监|专员|顾问|主管|科学家|讲师|教授|代理人|负责人|老师|培训师|运营师|规划师|管理师|首席信息官|首席技术官|首席数据官|首席安全官|首席运营官|CIO|CTO|CFO|COO|CMO|CEO|CPO|CDO|CXO)')
_TECH_ROLE_RE = re.compile(r'(算法|开发|数据|产品|测试|运维|前端|后端|全栈|安全|网络|架构|芯片|硬件|嵌入|量化|风控|建模|科研|销售|运营|市场|设计|人力|财务|法务|客服|提示词|prompt|智驾|自动驾驶|数字孪生|机器人|无人机)')
_AI_SIGNAL_RE = re.compile(r'(?i)(ai|nlp|cv|ml|dl|llm|aigc|大模型|智能体|agent|深度学习|机器学习|人工智能|rag|多模态|具身智能)')
# 前沿AI信号（真正的新岗位）：排除了传统AI如NLP/CV/ML/DL/深度学习
_FRONTIER_AI_RE = re.compile(r'(?i)(大模型|llm|智能体|agent|aigc|prompt|rag|具身智能|多模态大模型|'
    r'大语言模型|生成式|diffusion|stable\s*diffusion|gpt|claude|langchain|llamaindex|'
    r'autogpt|crewai|dify|coze|function\s*calling|tool\s*use|人形机器人)')

# 负向信号：有这些词 → 不像是岗位名（福利/招聘/地点/公司）
_NOISE_RE = re.compile(r'(校招|应届|往届|届|campus|管培|培训生|双休|法休|五险|一金|base|公积金|包吃|包住|下午茶|零食|团建|弹性|年假|年终|tob|toc|saas|独角兽|正编|外包|兼职|全职|远程|onsite|社招|计划|项目|方向职位|薪资open|待遇优|福利好|急聘|高薪|五险一金|带薪|年假)')
_COMPANY_RE = re.compile(r'(华为|腾讯|阿里|字节|百度|京东|美团|滴滴|快手|拼多多|网易|新浪|搜狐|'
                         r'tiktok|科大|讯飞|商汤|旷视|依图|第四范式|寒武纪|地平线|小马|文远|'
                         r'momenta|蔚来|小鹏|理想|比亚迪|小米|oppo|vivo|荣耀|中兴|新华三|'
                         r'浪潮|联想|360|奇安信|深信服|启明星辰|蚂蚁|米哈游|莉莉丝|叠纸|鹰角|'
                         r'趣加|沐瞳|友塔|三七|完美世界|西山居|畅游|盛趣|巨人)')
_DEPT_RE = re.compile(r'(研发部|技术部|产品部|设计部|测试部|数据部|运营部|市场部|销售部|财务部|人事部|行政部|事业部|事业群|研究院|实验室)')

# 公司名前缀（较长名优先→短名，用于 fused 公司名如"华为云计算AI…"）
_COMPANY_PREFIX_RE = re.compile(
    r'^(华为云计算|华为云|阿里云计算|阿里云|腾讯云计算|腾讯云|字节跳动|科大讯飞|'
    r'第四范式|小马智行|文远知行|新华三|启明星辰|完美世界|深信服|奇安信|'
    r'华为|阿里|腾讯|字节|百度|京东|美团|滴滴|快手|拼多多|网易|新浪|搜狐|'
    r'tiktok|商汤|旷视|依图|寒武纪|地平线|momenta|蔚来|小鹏|理想|比亚迪|'
    r'小米|oppo|vivo|荣耀|中兴|浪潮|联想|360|蚂蚁|米哈游|莉莉丝|叠纸|鹰角|'
    r'趣加|沐瞳|友塔|三七|西山居|畅游|盛趣|巨人)'
)

# 职级前缀（这些词若出现在开头→去掉）
_LEVEL_PREFIX_RE = re.compile(r'^(高级|资深|初级|见习|助理|副|代理|实习|管培)')

# 职级/管理后缀（这些词若出现在末尾→去掉）
_LEVEL_SUFFIX_RE = re.compile(r'(管培生|实习生|助理|培训生|见习生)$')

# 公司名-城市 后缀噪音（如"base大同双休法休"）
# 非贪婪匹配：保留第一个福利关键词之前的内容
# 英文关键词用 (?<![a-z]) 替代 \b（\b对汉字-英文边界无效）
_BENEFIT_SUFFIX_RE = re.compile(
    r'^(.+?)(?:'
    r'(?<![a-z])(?:base|tob|toc|saas)(?![a-z])'
    r'|双休|法休|五险|一金|公积金|包吃|包住|下午茶|零食|团建|弹性|年假|年终'
    r'|独角兽|正编|外包|兼职|全职|远程|社招'
    r').*$'
)


def _segment_score(seg):
    """对分隔后的一个段打分：正向高→更可能是岗位名"""
    score = 0
    # 正向
    if _JOB_SUFFIX_RE.search(seg):
        score += 20
    if _TECH_ROLE_RE.search(seg):
        score += 10
    if _AI_SIGNAL_RE.search(seg):
        score += 8
    if re.search(r'(?i)(python|java|golang|rust|c\+\+|scala|kotlin|swift|typescript|javascript)', seg):
        score += 6
    # 负向
    noise_count = len(_NOISE_RE.findall(seg))
    score -= noise_count * 20
    company_count = len(_COMPANY_RE.findall(seg))
    score -= company_count * 15
    if _DEPT_RE.search(seg):
        score -= 10
    # 长度
    if 4 <= len(seg) <= 22:
        score += 3
    elif len(seg) < 3:
        score -= 10
    elif len(seg) > 30:
        score -= 5
    # 纯英文/数字/下划线（无中文）通常不是中文岗位名
    if not re.search(r'[一-鿿]', seg):
        if re.match(r'^[a-z0-9_\-\.]+$', seg) and len(seg) < 20:
            score -= 8
    return score


def normalize_job_name(raw):
    """归一化岗位名：段打分+噪声过滤+概念合并

    流程：粗清→去福利后缀→按分隔符拆段→每段打分→取最佳段→段内清理→语义归一→大写恢复"""
    if not raw:
        return ''
    s = raw.strip()

    # ═══════════════ 第1轮：粗清理 ═══════════════
    # 去【】[] 等营销标签
    s = re.sub(r'[【\[][^】\]]*[】\]]', '', s)
    # 去所有括号内容（全角+半角），多轮处理嵌套
    for _ in range(3):
        new_s = re.sub(r'[（(][^)）]*[)）]', '', s)
        if new_s == s:
            break
        s = new_s
    # lowercase + 去空白
    s = s.lower().strip()
    s = re.sub(r'\s+', '', s)

    # ═══════════════ 第2轮：粗去福利/地点后缀 ═══════════════
    # 去掉 "base大同双休法休" / "双休+tob销售+SAAS独角兽" 类噪声
    stripped = _BENEFIT_SUFFIX_RE.sub(r'\1', s)
    # 只有剩余内容≥3字符且包含岗位信号才采纳（防止纯噪声串被截成碎片）
    if len(stripped) >= 3 and (_TECH_ROLE_RE.search(stripped) or _JOB_SUFFIX_RE.search(stripped) or _AI_SIGNAL_RE.search(stripped)):
        s = stripped
    # 去掉尾部 "/实习" "/全职" "/应届" "/社招" "/校招"
    s = re.sub(r'/[^\w一-鿿]{0,2}(实习|全职|应届|社招|校招|正式|兼职|外包)$', '', s)
    # 去掉尾部 "-campus-XXXX" / "-campus"
    s = re.sub(r'[-—]campus[-—]\d{4}$', '', s)
    s = re.sub(r'[-—]campus$', '', s)

    # ═══════════════ 第3轮：按分隔符拆段 → 打分 → 取最佳 ═══════════════
    # 分隔符: - — – / | ｜ . +（保护"C++"不被拆分）
    s = re.sub(r'(?i)c\+\+', 'CPLUSPLUS_PROTECT', s)
    segments = re.split(r'[-—–/|｜.+]+', s)
    segments = [seg.replace('CPLUSPLUS_PROTECT', 'C++') for seg in segments if seg]  # 去空+恢复C++

    if len(segments) >= 2:
        scored = [(seg, _segment_score(seg)) for seg in segments]
        best_seg, best_score = max(scored, key=lambda x: x[1])
        # 只有当最佳段得分>0才采用，否则保留原始（可能是特殊岗位名）
        if best_score > 0:
            s = best_seg
        # 否则保留全部（由后续LLM兜底）

    # ═══════════════ 第4轮：段内清理 ═══════════════
    # 去掉开头残留分隔符
    s = re.sub(r'^[-—–/|｜.]+', '', s)
    s = re.sub(r'[-—–/|｜.]+$', '', s)

    # 去掉公司名前缀（fused 如 "华为云计算AI..." → "AI..."）
    s = _COMPANY_PREFIX_RE.sub('', s)

    # 去掉职级前缀（高级/资深/初级/见习/助理/管培）
    s = _LEVEL_PREFIX_RE.sub('', s)

    # 去掉开头数字届数（如 "2027" "27届"）
    s = re.sub(r'^\d{2,4}届?', '', s)

    # 去掉残留 "方向" "岗" 后缀
    s = re.sub(r'(方向|岗)$', '', s)

    # 去掉残留 "-" 前后空白
    s = re.sub(r'^[-—]+', '', s)
    s = re.sub(r'[-—]+$', '', s)

    # ═══════════════ 第5轮：语义归一化 ═══════════════
    # aiagent / ai agent / agent → AI智能体（与"智能体"聚类合并）
    s = re.sub(r'(?i)^ai\s*agent', 'AI智能体', s)
    # 裸 "agent"（前面没有AI、后面不是英文字母→不截断"agentic"）→ AI智能体
    s = re.sub(r'(?i)^agent(?![a-z])', 'AI智能体', s)
    # 裸"智能体"（前面没有AI）→ 补AI前缀，确保三者合一
    s = re.sub(r'^(?!ai)智能体', 'AI智能体', s)

    # "实习生" → "工程师"（仅技术角色）
    # 大模型算法实习生 → 大模型算法工程师
    s = re.sub(
        r'(算法|开发|测试|数据|产品|设计|运维|安全|前端|后端|全栈|ai|nlp|cv|ml|'
        r'it|系统|网络|嵌入|硬件|芯片|结构|材料|电气|机械|光学|声学|仿真|'
        r'量化|风控|建模|分析|架构|科研|研究|大模型|深度学习|机器学习'
        r'|人工智能|自动驾驶|机器人|云计算|大数据)实习生$',
        r'\1工程师', s
    )
    # 去掉残留"实习生"（非技术前缀的情况）
    s = re.sub(r'实习生$', '', s)

    # 去掉职级/管理后缀（管培生/助理/培训生）— 放在实习生→工程师转换之后
    s = _LEVEL_SUFFIX_RE.sub('', s)

    # 去掉"高级"（无论位置 — 高级工程师/高级算法/Java高级开发）
    s = s.replace('高级', '')

    # 统一: 研发工程师/应用开发工程师/软件开发工程师 → 开发工程师
    s = re.sub(r'(应用开发|软件开发|研发)工程师', '开发工程师', s)
    # 应用工程师 → 开发工程师
    s = re.sub(r'应用工程师', '开发工程师', s)

    # 裸"AI工程师"补全 → "AI开发工程师"（AI不是独立工种，开发才是）
    s = re.sub(r'^ai工程师$', 'ai开发工程师', s)
    s = re.sub(r'^人工智能工程师$', '人工智能开发工程师', s)

    # ═══════════════ 第6轮：英文缩写/语言名大写恢复 ═══════════════
    for abbr in ['ai', 'nlp', 'cv', 'ml', 'dl', 'llm', 'aigc', 'rag', 'gpu', 'api', 'sdk', 'saas', 'paas']:
        s = re.sub(rf'(^|[^a-z]){abbr}([^a-z]|$)', rf'\1{abbr.upper()}\2', s)
    s = re.sub(r'(^|[^a-z])hr([^a-z]|$)', r'\1HR\2', s)
    s = re.sub(r'(^|[^a-z])it([^a-z]|$)', r'\1IT\2', s)
    # C-level 或其他全大写缩写
    for c_title in ['cio', 'cto', 'cfo', 'coo', 'cmo', 'ceo', 'cpo', 'cdo']:
        s = re.sub(rf'(^|[^a-z]){c_title}([^a-z]|$)', rf'\1{c_title.upper()}\2', s)

    # 编程语言/框架名大写
    s = re.sub(r'(^|[^a-z])(python|java|golang|rust|scala|kotlin|swift|ruby|perl|haskell|elixir)([^a-z]|$)',
               lambda m: m.group(1) + m.group(2).capitalize() + m.group(3), s)

    # 裸技术名词补全为岗位名
    s = re.sub(r'(?i)^python$', 'Python开发工程师', s)
    s = re.sub(r'(?i)^java$', 'Java开发工程师', s)
    s = re.sub(r'(?i)^golang$', 'Golang开发工程师', s)
    s = re.sub(r'(?i)^c\+\+$', 'C++开发工程师', s)
    s = re.sub(r'(?i)^(python|java|golang)工程师$', lambda m: m.group(1).capitalize() + '开发工程师', s)

    # ═══════════════ 第7轮："/" 分隔的复合名（兜底） ═══════════════
    if '/' in s:
        parts = [p.strip() for p in s.split('/')]
        best = max(parts, key=lambda p: (
            1 if _JOB_SUFFIX_RE.search(p) else 0,
            len(p)
        ))
        if best:
            s = best

    # ═══════════════ 第8轮：语义修复映射 ═══════════════
    # 技能名→岗位名（裸技能名不是岗位）
    IGN = re.IGNORECASE
    s = re.sub(r'^(机器学习|深度学习|强化学习|迁移学习|元学习|表示学习|对比学习|联邦学习)$', r'\1工程师', s, flags=IGN)
    s = re.sub(r'^(nlp算法|nlp|自然语言处理)$', r'NLP算法工程师', s, flags=IGN)
    s = re.sub(r'^(cv算法|cv|计算机视觉)$', r'CV算法工程师', s, flags=IGN)
    s = re.sub(r'^ai算法$', r'AI算法工程师', s, flags=IGN)
    s = re.sub(r'^aigc算法$', r'AIGC算法工程师', s, flags=IGN)
    s = re.sub(r'^aigc$', r'AIGC算法工程师', s, flags=IGN)
    s = re.sub(r'^ai智能体$', r'AI智能体开发工程师', s, flags=IGN)

    # 专家→工程师（同一聚类）
    s = re.sub(r'^(ai大模型|ai|大模型|aigc|nlp|cv|搜推|推荐|风控|语音|视觉|图像|)算法专家$', r'\1算法工程师', s, flags=IGN)
    s = re.sub(r'^算法专家$', r'算法工程师', s)
    s = re.sub(r'^ai(大模型|)算法专家$', r'AI算法工程师', s, flags=IGN)

    # 提示词→Prompt（AI前缀后加空格）
    s = re.sub(r'ai提示词', 'AI Prompt', s, flags=IGN)
    s = re.sub(r'提示词', 'Prompt', s)

    # AI研发→AI研发工程师
    s = re.sub(r'^ai研发$', r'AI研发工程师', s, flags=IGN)
    # 研发工程师/研发→开发工程师（已有规则，此处兜底裸"研发"）
    s = re.sub(r'^研发$', r'开发工程师', s)
    s = re.sub(r'^ai应用研发$', r'AI应用开发工程师', s, flags=IGN)

    # 算法负责人→算法团队负责人
    s = re.sub(r'^算法负责人$', r'算法团队负责人', s)

    # 智驾→自动驾驶
    s = re.sub(r'^智驾工程师$', r'自动驾驶算法工程师', s)
    s = re.sub(r'^智驾算法工程师$', r'自动驾驶算法工程师', s)
    s = re.sub(r'^智驾', r'自动驾驶', s)

    # AI+人工智能冗余→AI（如"AI人工智能算法工程师"→"AI算法工程师"）
    s = re.sub(r'ai人工智能', 'AI', s, flags=IGN)
    s = re.sub(r'人工智能ai', 'AI', s, flags=IGN)

    # 裸"工程师" → 如果含AI信号就补全为AI开发工程师
    if s == '工程师':
        s = 'AI开发工程师'

    # AI方向职位 → 这不是岗位名，是个类别标题
    if 'AI方向职位' in s:
        s = s.replace('AI方向职位', 'AI开发工程师')

    # 助理工程师→工程师
    s = re.sub(r'^AI开发助理工程师$', 'AI开发工程师', s)

    # 尾部残留 "+" "+五险一金" "+带薪" 等噪声
    s = re.sub(r'\+.*$', '', s)

    # ═══════════════ 第9轮：无岗位后缀的AI/技术名→自动补全 ═══════════════
    # 如果有AI信号或技术角色，但没有岗位后缀，自动补全
    has_ai = bool(_AI_SIGNAL_RE.search(s))
    has_tech = bool(_TECH_ROLE_RE.search(s))
    has_suffix = bool(_JOB_SUFFIX_RE.search(s))
    if (has_ai or has_tech) and not has_suffix:
        # 已经是完整表述的例外：Prompt/Agent/RAG/AIGC/LLM/NLP/CV/ML/DL 单独出现
        if re.match(r'^(prompt|agent|rag|aigc|llm|nlp|cv|ml|dl)$', s, flags=IGN):
            s = s.upper() + '开发工程师' if len(s) <= 4 else s + '工程师'
        elif s.endswith('开发'):
            s = s + '工程师'       # "AI智能体开发"→"AI智能体开发工程师"
        elif s.endswith('算法'):
            s = s + '工程师'       # "NLP算法"→"NLP算法工程师"（兜底，上面已处理大部分）
        elif s.endswith('测试'):
            s = s + '工程师'
        elif s.endswith('运维'):
            s = s + '工程师'
        elif s.endswith('研发'):
            s = s + '工程师'
        elif s.endswith('设计'):
            s = s + '师'
        elif s.endswith('分析'):
            s = s + '师'
        else:
            s = s + '开发工程师'   # 兜底补全

    # ═══════════════ 最终清理 ═══════════════
    s = s.strip()
    s = re.sub(r'^[-—–/|｜.]+', '', s)
    s = re.sub(r'[-—–/|｜.]+$', '', s)
    # 去连续重复词（"开发开发"→"开发"、"智能体智能体"→"智能体"）— 同词重复才去
    s = re.sub(r'(开发|测试|算法|设计|运维|智能体|agent|工程师|架构)\1', r'\1', s, flags=IGN)
    # 去残留招聘噪声（循环，处理"工程师招聘"→"工程师招"→"工程师"）
    for _ in range(3):
        prev = s
        s = re.sub(r'[岗招聘急校]$', '', s)
        s = re.sub(r'招聘$', '', s)
        if s == prev:
            break
    # 去纯英文连写（"senioraiengineer"→删除）
    s = re.sub(r'[a-z]{15,}$', '', s, flags=IGN)
    # 全噪声检测：无AI信号、无技术角色、无岗位后缀 → 拒绝
    if not (_AI_SIGNAL_RE.search(s) or _TECH_ROLE_RE.search(s) or _JOB_SUFFIX_RE.search(s)):
        return ''
    if len(s) < 3:
        return ''
    return s


# ============================================================
#  技能解析
# ============================================================

def parse_skill_list(s):
    """解析 【技能|类别|级别】；【...】 格式 → [{'skill','category','level'}]"""
    if not s or pd.isna(s):
        return []
    skills = []
    for m in re.finditer(r'【(.+?)】', str(s)):
        parts = m.group(1).split('｜')
        if len(parts) >= 1:
            skills.append({
                'skill': parts[0].strip(),
                'category': parts[1].strip() if len(parts) >= 2 else '未知',
                'level': parts[2].strip() if len(parts) >= 3 else '未标注',
            })
    return skills


# ============================================================
#  职责提取
# ============================================================

# ── 职责句清洗：噪声模式（薪资、福利、公司介绍等） ──
_DUTY_NOISE_PATTERNS = [
    # 薪资
    r'薪[资酬].{0,10}([0-9]+[kKwW万]|面议|可谈|\d+[-~]\d+)',
    # 福利
    r'五险|公积金|社保|年终奖|绩效奖|股票期权|弹性工作|双休|团建|下午茶|零食',
    # 公司介绍（成立于/融资/上市等）
    r'(公司|企业|我们).{0,5}(成立于|是一家|总部|融资|上市|纳斯达克|A股)',
    # 招聘套话
    r'欢迎.{0,5}(加入|投递|来撩)|急聘|高薪|待遇优|福利好',
    # 公司问候语/自我介绍（"尊敬的求职者您好，我们是XX公司"）
    r'尊敬的.{0,5}(求职者|应聘者|面试者)|您好.{0,10}(我们是|我是)',
    r'我们是.{0,20}(公司|科技|科技公司|团队|部门|技术部门|深耕|专注|一家)',
    # 公司愿景/产品野心（不是个人职责）
    r'我们希望|我们的目标|我们的愿景|我们的使命',
    r'致力于打造|致力于成为|致力于推动',
    # 产品营销口号
    r'降低.{0,5}(创造|制作|开发|硬件)?.{0,3}门槛|享受.{0,5}(心流|创造)',
    r'打造.{0,5}(属于|一款|一个).{0,5}(的|之)',
    # 任职要求/资格（不是职责） — 用 .*? 而非 .{0,N}，因为中英文混排不可预测
    r'(以上|年以上|年以上相关|具备|工作经验).*?(经验|工作经历|项目经验|实战经验)',
    r'(有|具有|具备).*?(经验|优先|者优先|加分)',
    r'(经验|工作经历|项目经历).*?(优先|者优先|加分)',
    r'(精通|熟练掌握|熟练使用|熟悉).*?(优先|者优先|加分|更佳|框架|工具|编程|开发|语言|平台)',
    r'^(熟练掌握|熟练使用|精通|熟悉).{0,40}(框架|工具|开发|编程|语言|系统|平台|模型)',
    r'(本科|硕士|博士|学历|统招|全日制)',
    r'(硬性|加分项|优先条件|我们希望你)',
    # 岗位标题/名称混入职责
    r'^岗位[：:\s]|岗位名称|职位名称',
    # 联系方式/投递
    r'(投递|发送简历|联系邮箱|微信|电话).{0,20}',
    # 纯关键词/技术栈堆砌（缺少中文动词，如 "DEEPSEEK GPT CLAUDE AI系统 Python..."）
    r'^[a-zA-Z\s\+#\.\-_]{30,}$',
    # 短句以"精通/熟练掌握/熟悉"开头且无职责动词 → 技能陈述，非职责
    r'^(精通|熟练掌握|熟练使用|熟悉)(?!.*(负责|参与|设计|开发|优化|实现|构建|搭建|推动|落地|迭代|撰写|制定|规划|管理|协调|分析|调研|探索|跟踪|维护|部署|测试|交付))',

    # 工作经验开头
    r'^工作经验[：:]',
    # 公司营销口号/产品广告（不是职责，如"打开高手脉，瓜分百亿"）
    r'(超级品牌|瓜分百亿|八大理由|股权激励|秉承.*使命|全球先进|高达.*亿|加入.*理由|惠及全民)',
    r'(轻轻松松|不要错过|绝佳机会|千载难逢|颠覆|革命性.{0,5}(产品|技术))',
    r'高手脉',  # 特定垃圾JD公司名，职责中嵌入公司广告
    # 课程/培训安排（不是职责，如"6月22日-7月11日，共三周"）
    r'(\d+月\d+日[至\-~]\d+月\d+日|共\d+周|每周[一二三四五六日]|授课|实训项目)',
    r'(每天预计|晚上.*答疑|周日休息|周一至周六)',
    # 销售/招聘话术
    r'(你将(免费|掌握|学会|获得|拥有|体验))',
    r'(急需.{0,5}人|高薪急招|急招)',
    # 薪资/课酬
    r'(课酬|课时费|薪资范围|薪酬|月薪)',
]

def _clean_duty_sentence(s):
    """清洗单条职责句：去标记、去噪声、规范化"""
    if not s or len(s) < 6:
        return ''
    # 去掉 Markdown 标记：### / ## / # 标题、**加粗**、__下划线__
    s = re.sub(r'^[\s]*#{1,3}\s*', '', s)
    s = re.sub(r'\*\*([^*]+)\*\*', r'\1', s)
    s = re.sub(r'__([^_]+)__', r'\1', s)
    # 去掉残留的裸 * 和 # 符号（不在词中的）
    s = re.sub(r'^[\s]*[*#]+\s*', '', s)
    # 去掉【】标签（JD正文中的公司口号/编辑标记，不是职责）
    s = re.sub(r'【[^】]*】', '', s)
    # 去孤儿括号：单独的 "】" 或 "】" 开头
    s = re.sub(r'^[\s]*[】\]][\s]*', '', s)
    s = re.sub(r'[】【\[\]]+', ' ', s)
    # 去掉列表标记前缀（含 * 星号）
    s = re.sub(r'^[\s]*[-—•·▪▸►●○◆◇▪▹▪◦∙﹫*]+[\s]*', '', s)
    s = re.sub(r'^[\s]*[\d]+[、.．)\s]*', '', s)
    # 去掉中文序号：一、二、三、... 十、
    s = re.sub(r'^[\s]*[一二三四五六七八九十]+[、．]\s*', '', s)
    s = re.sub(r'^[\s]*[①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳]+[\s]*', '', s)
    s = re.sub(r'^[\s]*[（(][\d]+[)）][\s]*', '', s)
    # 去掉段落标题（含前置修饰，循环处理"xxx 任职要求：一、基本任职要求" 嵌套情形）
    for _ in range(3):
        prev = s
        s = re.sub(r'^.{0,50}(岗位职责|工作内容|职位描述|工作职责|岗位描述|岗位要求|任职要求|任职资格|岗位内容|基本任职要求|任职条件|岗位概述|技术要求)[：:]?\s*', '', s)
        # 再次去掉序号（可能在标题剥离后才暴露）
        s = re.sub(r'^[\s]*[-—•·▪▸►●○◆◇▪▹▪◦∙﹫*]+[\s]*', '', s)
        s = re.sub(r'^[\s]*[\d]+[、.．)\s]*', '', s)
        s = re.sub(r'^[\s]*[一二三四五六七八九十]+[、．]\s*', '', s)
        if s == prev:
            break
    # 去掉残留的裸标题关键词
    s = re.sub(r'^(职责|要求|资格|描述|工作|任务)[：:]\s*', '', s)
    # 去括号中的百分比标识："AI大模型应用（15%）" → "AI大模型应用"
    s = re.sub(r'[（(]\s*\d{1,3}\s*%\s*[）)]', '', s)
    # 去"方向X：" / "方向四：" 类型的段落标题
    s = re.sub(r'^[\s]*方向[一二三四五六七八九十\d]+[：:]\s*[^\s]{0,20}方向\s*[-—–]?\s*', '', s)
    s = re.sub(r'^[\s]*方向[一二三四五六七八九十\d]+[：:]\s*', '', s)
    # 去掉括号标注标记："（核心）"、"（重点）"、"（必选）" 等
    s = re.sub(r'[（(]\s*(核心|重点|必选|必须|关键|主要|加分|优先|必备)\s*[）)]', '', s)
    # 去掉尾部残留数字（如 "AI Agent架构设计 1" → "AI Agent架构设计"）
    s = re.sub(r'\s+\d+\s*$', '', s)
    # 过滤：看起来像岗位名而非职责的句子（如"算法研究员（多模态/视觉/图像/大模型方向）"）
    if re.match(r'^[\w一-鿿]{2,20}工程师[（(].{0,30}[）)]$', s):
        return ''
    if re.match(r'^[\w一-鿿]{2,15}研究员[（(].{0,30}[）)]$', s):
        return ''
    # 过滤：截断句（以英文单词片段结尾，如"Fu" "MC"等2-3字母）
    if re.search(r'[，,、]\s*[A-Za-z]{1,3}$', s) and len(s) < 60:
        return ''
    # 过滤纯噪声行
    for pat in _DUTY_NOISE_PATTERNS:
        if re.search(pat, s):
            return ''
    # 过滤：清洗后只剩纯标点/空白/过短的
    if len(s.strip('，。；;：:、, \t\r\n（）()[]【】""''！!？?')) < 5:
        return ''
    # 规范化：去掉行尾无意义字符
    s = s.strip('，。；;：:、, \t\r\n')
    return s

def parse_duties(text):
    """从JD原文提取职责句子（含清洗）"""
    if not text or len(text) < 20:
        return []
    # 段落级职责区段匹配
    duty_patterns = [
        r'岗位职责[：:](.*?)(?=岗位要求|任职要求|任职资格|岗位条件|$)',
        r'工作(内容|职责|描述)[：:](.*?)(?=岗位要求|任职要求|任职资格|岗位条件|$)',
        r'职位描述[：:](.*?)(?=岗位要求|任职要求|任职资格|岗位条件|$)',
    ]
    duty_text = ''
    for pat in duty_patterns:
        m = re.search(pat, text, re.DOTALL)
        if m:
            duty_text = m.group(0)
            break

    if not duty_text:
        # 无明确标题，取全文前60%
        duty_text = text[:int(len(text) * 0.6)]

    # 拆句子 + 清洗
    raw_sentences = re.split(r'[。；;.\n]', duty_text)
    sentences = []
    for s in raw_sentences:
        cleaned = _clean_duty_sentence(s.strip())
        if cleaned and 8 < len(cleaned) < 150:
            sentences.append(cleaned)

    # 合并过短的相邻句（正向：短句并入下一句；反向：短句并入上一句）
    # 第一遍：正向合并（短句→下一句）
    i = 0
    while i < len(sentences) - 1:
        if len(sentences[i]) < 15:
            sentences[i+1] = sentences[i] + '；' + sentences[i+1]
            sentences[i] = ''
        i += 1
    sentences = [s for s in sentences if s]
    # 第二遍：反向合并（短句→上一句，向后兜底）
    merged = []
    for cur in sentences:
        if len(cur) < 15 and merged:
            merged[-1] = merged[-1] + '；' + cur
        else:
            merged.append(cur)
    return merged


def score_duties(sentences, ai_signal_set, cluster_duty_counter, N):
    """用AI信号+簇内频率评分，返回Top 5-8条"""
    # 先去完全重复（同一个句子出现多次只在簇内保留一次）
    seen_exact = set()
    unique_sentences = []
    for sent in sentences:
        key = sent.strip().lower()
        if key not in seen_exact:
            seen_exact.add(key)
            unique_sentences.append(sent)
    sentences = unique_sentences

    scored = []
    for sent in sentences:
        if len(sent) < 10:
            continue
        # AI信号密度
        ai_hits = sum(1 for sig in ai_signal_set if sig.lower() in sent.lower())
        ai_density = ai_hits / max(len(sent), 1)
        # 簇内频率分（bigram近似）
        freq_score = 0
        for i in range(len(sent) - 1):
            bigram = sent[i:i+2]
            freq_score += cluster_duty_counter.get(bigram, 0)
        freq_score /= max(len(sent), 1)
        # 综合分
        score = 0.4 * ai_density + 0.3 * freq_score + 0.3 * (1.0 / (len(sentences) + 1))
        scored.append((score, sent))

    scored.sort(key=lambda x: -x[0])

    # 去重
    result = []
    for _, sent in scored:
        dup = False
        for existing in result:
            if SequenceMatcher(None, sent, existing).ratio() > 0.7:
                dup = True
                break
        if not dup:
            result.append(sent)
        if len(result) >= 8:
            break
    return result[:8]


def extract_ai_signals_from_text(text):
    """从文本提取AI信号"""
    if not text:
        return set()
    return {sig for sig in AI_SIGNALS if sig.lower() in text.lower()}


# ============================================================
#  行业推断
# ============================================================

def infer_industry(company_name, work_area='', jd_text='', job_name=''):
    """从岗位名+职责+公司名推断行业（岗位名权重最高，职责次之，公司名兜底）"""
    jd_full = (jd_text or '')[:2000]
    company_text = f'{company_name} {work_area}'
    name_text = job_name or ''

    # 收集所有匹配的行业及其关键词命中数
    # 权重: 岗位名×5 > JD文本×2 > 公司名×1
    scores = {}
    for pattern, industry in INDUSTRY_MAP:
        nm_matches = len(re.findall(pattern, name_text)) if name_text else 0
        jd_matches = len(re.findall(pattern, jd_full)) if jd_full else 0
        co_matches = len(re.findall(pattern, company_text))
        if nm_matches > 0 or jd_matches > 0 or co_matches > 0:
            scores[industry] = scores.get(industry, 0) + nm_matches * 5 + jd_matches * 2 + co_matches

    if scores:
        return max(scores, key=scores.get)

    # JD关键词兜底
    jd_lower = jd_full.lower()
    for industry, keywords in INDUSTRY_JD_KEYWORDS.items():
        for kw in keywords:
            if kw.lower() in jd_lower:
                return industry
    return '通用信息技术'


# ============================================================
#  LLM 精炼（可选兜底）
# ============================================================

def call_llm_api(prompt, api_key, temperature=0.4):
    """调用 DeepSeek API（urllib，零外部依赖）"""
    import urllib.request
    import urllib.error

    payload = {
        'model': LLM_CONFIG['model'],
        'messages': [
            {'role': 'system',
             'content': '你是一位专业的HR分析师和岗位定义专家。你的输出必须严格基于提供的源数据，不得编造任何技能名、公司名或统计数据。'},
            {'role': 'user', 'content': prompt},
        ],
        'temperature': temperature,
        'max_tokens': LLM_CONFIG['max_tokens'],
    }

    data = json.dumps(payload).encode('utf-8')
    req = urllib.request.Request(
        f'{LLM_CONFIG["api_base"].rstrip("/")}/chat/completions',
        data=data,
        headers={
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {api_key}',
        },
    )

    for attempt in range(5):
        try:
            with urllib.request.urlopen(req, context=_SSL_CONTEXT, timeout=90) as resp:
                result = json.loads(resp.read().decode('utf-8'))
                return result['choices'][0]['message']['content']
        except urllib.error.HTTPError as e:
            body = e.read().decode('utf-8')[:300] if e.fp else ''
            if e.code == 429:
                time.sleep(2 ** attempt)
            elif e.code >= 500:
                time.sleep(2 ** attempt)
            else:
                print(f'  [LLM] API错误 {e.code}: {body}')
                return None
        except Exception as e:
            print(f'  [LLM] 网络错误: {e}')
            if attempt < 4:
                time.sleep(2 ** attempt)
    return None


def build_llm_prompt(defn):
    """基于规则定义构建LLM精炼prompt"""
    t = defn['source_traceability']
    raw_duties = defn['核心职责']
    req_skills = [s['skill'] for s in defn['必备技能']]
    bonus_skills = [s['skill'] for s in defn['加分技能']]
    industries = [i['industry'] for i in defn['典型行业应用场景']]

    prompt = f"""请基于以下源数据，为岗位"{defn['岗位名称']}"生成精炼的岗位定义。

【源数据】
- 岗位名称: {defn['岗位名称']}
- 别称: {', '.join(defn['别称'][:5])}
- 样本量: {t['total_jds']}条JD, {t['unique_companies']}家公司
- 代表性公司: {', '.join(t['top_companies'][:5])}
- 代表城市: {', '.join(t['top_cities'][:5])}

【原始职责描述（从{len(raw_duties)}条规则提取，可能含切片/噪声）】
{chr(10).join(f'- {d[:150]}' for d in raw_duties[:12])}

【高频必备技能（来自{t['total_jds']}条JD统计）】
{', '.join(req_skills[:8])}

【加分技能】
{', '.join(bonus_skills[:6])}

【规则推断行业】
{', '.join(industries[:5])}

请输出JSON格式（只输出JSON，不要其他文字）：
{{
  "核心职责": [
    "职责1（≤30字，凝练而非照抄，但必须能从源职责回溯）",
    ...
  ],
  "典型行业应用场景": [
    {{"industry": "行业名", "rationale": "一句话理由（必须引用源数据中的公司或JD内容）"}},
    ...
  ]
}}

要求：
1. 核心职责5-8条，每条≤35字，用"负责/参与/设计/开发/优化/推动/搭建/落地/调研/迭代"等动词开头
2. 【禁止】输出截断的句子（如"有大模型微调、Fu"）、岗位名称（如"算法研究员（多模态方向）"）、
   任职要求（如"年以上经验"）、公司广告语、关键词堆砌（如"LLM数据生成，模型评测"）
3. 遇到源职责中明显噪声/截断/广告/技能陈述时，改成职责口吻、补全语义，但不得编造技能名
4. 行业场景3-5个，必须基于源数据中的公司名推断，不得凭空添加行业
5. 严禁编造源数据中不存在的技能、公司、数字
6. 源职责中的技能陈述（如"熟练掌握Python，精通Pytorch"）不属于核心职责，请改写为工作任务描述
   或跳过，提取真正描述"做什么"的句子
7. 只输出JSON，不要markdown代码块标记"""
    return prompt


def verify_llm_duties(llm_duties, source_duties, source_skills, source_companies):
    """校验LLM输出：检测幻觉（编造技能/公司/数字）

    四重验证：
    1. 技能词命中：LLM凝练的职责是否引用了源数据中的技能词
    2. 字符3-gram重叠：是否与源职责文本有基本字符交集（阈值8%）
    3. 公司名编造检测
    4. 禁止词检测：截断片段/岗位名/广告语
    """
    warnings = []
    source_text = ' '.join(source_duties).lower()
    all_source_skills = set(s.lower() for s in source_skills)
    all_source_companies = set(c.lower() for c in source_companies)

    # 禁止模式（LLM不应输出的内容）
    FORBIDDEN_PATTERNS = [
        (r'[，,]\s*[A-Za-z]{1,3}$', '截断片段（以1-3字母结尾）'),
        (r'工程师[（(].{0,20}[）)]$', '岗位名当职责'),
        (r'研究员[（(].{0,20}[）)]$', '岗位名当职责'),
        (r'.{0,5}(任职要求|岗位要求|工作经验|学历要求)', '段落标题泄露'),
        (r'(超级品牌|瓜分百亿|八大理由|股权激励|惠及全民)', '公司广告语'),
    ]

    for duty in llm_duties:
        if len(duty) < 5:
            warnings.append(f'职责过短: {duty}')
            continue

        duty_lower = duty.lower()

        # 检查0：禁止模式检测
        forbidden_hit = False
        for pat, desc in FORBIDDEN_PATTERNS:
            if re.search(pat, duty):
                warnings.append(f'禁止内容({desc}): {duty[:60]}...')
                forbidden_hit = True
                break
        if forbidden_hit:
            continue

        # 检查1：技能词命中（LLM凝练后仍应包含1个以上源技能词）
        skill_hits = sum(1 for sk in all_source_skills if len(sk) >= 2 and sk in duty_lower)
        # 检查2：字符3-gram重叠（阈值8%，凝练要求提高）
        trigrams = [duty[i:i+3] for i in range(len(duty)-2)]
        char_hit = sum(1 for t in trigrams if t in source_text)
        char_ratio = char_hit / max(len(trigrams), 1)

        # 双通道判断：技能命中≥1 或 字符重叠≥8% 即通过
        if skill_hits >= 1 or char_ratio >= 0.08:
            continue  # 通过
        else:
            warnings.append(f'职责与源数据关联弱(skill_hits={skill_hits}, char_ratio={char_ratio:.2f}): {duty[:60]}...')

        # 检查3：公司名编造检测（如果职责中出现公司名，必须是源数据中的）
        for company in all_source_companies:
            if len(company) >= 5 and company in duty_lower:
                # 公司名在源数据中 → OK
                break
        else:
            # 检查是否有未知公司名（中文字符≥4的专有名词）
            unknown_company = re.findall(r'[一-鿿]{4,10}(?:公司|科技|集团|有限|股份)', duty)
            if unknown_company:
                for uc in unknown_company:
                    if uc.lower() not in all_source_companies:
                        warnings.append(f'疑似编造公司名: {uc} in {duty[:60]}...')

    return warnings


def refine_with_llm(defn, api_key):
    """对单条定义做LLM精炼，返回精炼后的字段（失败则返回None）"""
    # 预处理：清理规则提取中残留的明显噪声（给LLM更干净的输入）
    cleaned_duties = []
    for d in defn['核心职责']:
        d = re.sub(r'[】【\[\]]+', ' ', d).strip()
        d = re.sub(r'\s+', ' ', d)
        if len(d) >= 8:
            cleaned_duties.append(d)
    defn['核心职责'] = cleaned_duties

    prompt = build_llm_prompt(defn)
    raw = call_llm_api(prompt, api_key, temperature=0.4)

    if not raw:
        print('[API失败]', end=' ')
        return None

    # 解析JSON（多层容错）
    raw = raw.strip()
    result = None

    # 策略1：直接解析
    try:
        # 清理markdown包裹
        if raw.startswith('```'):
            raw = re.sub(r'^```\w*\s*', '', raw)
            raw = re.sub(r'\s*```$', '', raw)
        result = json.loads(raw)
    except json.JSONDecodeError:
        pass

    # 策略2：正则提取最外层{}（非贪婪→贪婪逐步降级）
    if result is None:
        for pattern in [r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', r'\{[\s\S]*\}']:
            m = re.search(pattern, raw)
            if m:
                try:
                    result = json.loads(m.group(0))
                    break
                except json.JSONDecodeError:
                    continue

    # 策略3：修复常见JSON问题后重试
    if result is None:
        try:
            fixed = raw
            # 去掉尾部多余逗号
            fixed = re.sub(r',\s*([}\]])', r'\1', fixed)
            # 单引号→双引号（仅key和简单value）
            # 中文引号→去掉
            fixed = fixed.replace('“', '"').replace('”', '"')
            m = re.search(r'\{[\s\S]*\}', fixed)
            if m:
                result = json.loads(m.group(0))
        except json.JSONDecodeError:
            pass

    if result is None:
        print('[JSON解析失败]', end=' ')
        return None

    # 校验
    source_skills = [s['skill'] for s in defn['必备技能']] + [s['skill'] for s in defn['加分技能']]
    source_companies = defn['source_traceability']['top_companies']
    warnings = verify_llm_duties(
        result.get('核心职责', []),
        defn['核心职责'],
        source_skills,
        source_companies
    )

    # 拒绝阈值：超过70%职责有告警才拒绝（原50%过严）
    total_duties = len(result.get('核心职责', [1]))
    if total_duties > 0 and len(warnings) > total_duties * 0.7:
        print(f'[幻觉拒绝:{len(warnings)}/{total_duties}]', end=' ')
        return None

    # 少量告警不拒绝，只记录
    if warnings:
        print(f'[通过({len(warnings)}条弱关联)]', end=' ')

    return {
        '核心职责': result.get('核心职责', defn['核心职责']),
        '典型行业应用场景': result.get('典型行业应用场景', defn['典型行业应用场景']),
        'llm_warnings': warnings,
    }


def merge_similar_definitions(definitions):
    """合并同义不同名的岗位定义（名称高度相似→归入大簇）

    合并条件（满足条件1+2，且满足条件3A或3B任一）：
    1. 名称子串关系：ni in nj 或 nj in ni（差≤10字符）
    2. 核心角色一致：提取末词（工程师/研究员/专家…）必须相同
    3A. 大小簇合并：小簇(≤15 JD) 并入 大簇(≥15 JD)
    3B. 职责AI对齐：小簇职责含AI信号 → 可并入AI前缀的大簇（允许更大size差）"""
    merged = []
    used = [False] * len(definitions)
    for i, di in enumerate(definitions):
        if used[i]:
            continue
        ni = di['岗位名称']
        jds_i = di['source_traceability']['total_jds']
        cluster = [i]
        for j, dj in enumerate(definitions):
            if i == j or used[j]:
                continue
            nj = dj['岗位名称']
            jds_j = dj['source_traceability']['total_jds']
            # 条件1：短名是长名的严格子串（差≤8字符）
            substring_match = ((ni in nj or nj in ni) and abs(len(ni) - len(nj)) <= 8)
            # 条件2：末词一致（核心角色：工程师/研究员/专家/经理等）
            suffix_i = re.search(r'(工程师|研究员|专家|分析师|设计师|架构师|经理|总监|科学家|讲师|老师)$', ni)
            suffix_j = re.search(r'(工程师|研究员|专家|分析师|设计师|架构师|经理|总监|科学家|讲师|老师)$', nj)
            same_suffix = suffix_i and suffix_j and suffix_i.group(1) == suffix_j.group(1)
            if not (substring_match and same_suffix):
                continue
            # 条件3A：小簇≤5 JD且大簇≥20 JD（只合并极端不对称）
            big_small_strict = (jds_i >= 20 and jds_j <= 5) or (jds_j >= 20 and jds_i <= 5)
            # 条件3B：职责AI对齐 → 扩大阈值（小簇≤20 JD，大簇≥10 JD，且小簇职责含AI信号）
            small_j, large_j = (j, i) if jds_i > jds_j else (i, j)
            small_jds = dj['source_traceability']['total_jds'] if small_j == j else di['source_traceability']['total_jds']
            large_jds = di['source_traceability']['total_jds'] if large_j == i else dj['source_traceability']['total_jds']
            small_duties = ' '.join(dj.get('核心职责', []) if small_j == j else di.get('核心职责', []))
            ai_aligned = (
                small_jds <= 20 and large_jds >= 10 and
                _AI_SIGNAL_RE.search(small_duties) and
                _AI_SIGNAL_RE.search(ni if large_j == i else nj)
            )
            if big_small_strict or ai_aligned:
                cluster.append(j)
                used[j] = True
        if len(cluster) > 1:
            # 合并：取大簇名称为主，小簇JD全部并入
            main_idx = max(cluster, key=lambda x: definitions[x]['source_traceability']['total_jds'])
            main = definitions[main_idx]
            for k in cluster:
                if k == main_idx:
                    continue
                dk = definitions[k]
                # 合并核心职责（去重）
                existing_duties = set(main['核心职责'])
                for duty in dk.get('核心职责', []):
                    if duty not in existing_duties:
                        main['核心职责'].append(duty)
                        existing_duties.add(duty)
                # 合并必备技能（重新统计）
                for sk in dk.get('必备技能', []):
                    found = False
                    for ms in main['必备技能']:
                        if ms['skill'] == sk['skill']:
                            ms['support_pct'] = max(ms['support_pct'], sk['support_pct'])
                            found = True
                            break
                    if not found:
                        main['必备技能'].append(sk)
                # 合并加分技能
                for sk in dk.get('加分技能', []):
                    found = False
                    for ms in main.get('加分技能', []):
                        if ms['skill'] == sk['skill']:
                            found = True
                            break
                    if not found:
                        main.setdefault('加分技能', []).append(sk)
                # 合并别称
                for alias in dk.get('别称', []):
                    if alias not in main['别称']:
                        main['别称'].append(alias)
                # 更新溯源
                main['source_traceability']['total_jds'] += dk['source_traceability']['total_jds']
                main['source_traceability']['unique_companies'] += dk['source_traceability']['unique_companies']
            # 合并后重新截断核心职责到Top 8
            if len(main['核心职责']) > 8:
                main['核心职责'] = main['核心职责'][:8]
            main['必备技能'] = main['必备技能'][:8]
            if '加分技能' in main:
                main['加分技能'] = main['加分技能'][:8]
            used[main_idx] = True
            merged.append(main)
        else:
            merged.append(di)
    return merged


# ============================================================
#  主逻辑
# ============================================================

def load_new_jobs(input_dir):
    """加载所有已标注CSV，筛选新岗位候选"""
    files = list(set(
        glob.glob(os.path.join(input_dir, '【已标注】*.csv')) +
        glob.glob(os.path.join(input_dir, '*.csv'))
    ))
    if not files:
        files = list(set(glob.glob(os.path.join(input_dir, '**', '*.csv'), recursive=True)))

    rows = []
    for f in files:
        try:
            df = pd.read_csv(f)
            for _, row in df.iterrows():
                if str(row.get('是否新岗位候选', '')).strip() == '是':
                    rows.append(row.to_dict())
        except Exception as e:
            print(f'  [WARN] 无法读取 {f}: {e}')

    print(f'[LOAD] {len(files)} 个文件, {len(rows)} 条新岗位JD')
    return rows


def group_and_generate(rows, timeline_report_path=None, use_llm=False, llm_max=999):
    """按归一化岗位名分组，每组生成一份定义"""
    # 分组
    groups = defaultdict(list)
    for r in rows:
        raw = str(r.get('job_name', '')).strip()
        norm = normalize_job_name(raw)
        if not norm:
            continue
        groups[norm].append(r)

    print(f'[GROUP] {len(groups)} 个岗位类型')

    # 加载时间轴数据（可选）
    timeline_index = {}
    if timeline_report_path and os.path.exists(timeline_report_path):
        try:
            with open(timeline_report_path, 'r', encoding='utf-8') as f:
                cap = json.load(f)
            for rep in cap.get('reports', []):
                timeline_index[rep['job']] = rep.get('update_summary', '')
            print(f'[TIMELINE] {len(timeline_index)} 条能力趋势已加载')
        except Exception:
            pass

    # LLM API Key（优先环境变量 → 配置文件 → 无key则回退规则）
    api_key = None
    if use_llm:
        api_key = os.environ.get('DEEPSEEK_API_KEY', '') or LLM_CONFIG.get('api_key', '')
        if not api_key:
            print('[LLM] DEEPSEEK_API_KEY 未设置，回退纯规则模式')
            use_llm = False

    definitions = []
    llm_refined = 0
    for cluster_id, (norm_name, jds) in enumerate(sorted(groups.items(), key=lambda x: -len(x[1])), 1):
        n = len(jds)
        if n < 2:
            continue  # 单条JD不足以"定义"一个岗位

        defn = build_definition(cluster_id, norm_name, jds, timeline_index)

        # LLM精炼：全部岗位
        if use_llm and llm_refined < llm_max:
            try:
                print(f'  [LLM] 精炼 #{cluster_id} {defn["岗位名称"]} ({n}JDs)...', end=' ')
            except UnicodeEncodeError:
                print(f'  [LLM] 精炼 #{cluster_id} (name#{cluster_id}) ({n}JDs)...', end=' ')
            refined = refine_with_llm(defn, api_key)
            if refined:
                defn['核心职责'] = refined['核心职责']
                defn['典型行业应用场景'] = refined['典型行业应用场景']
                defn['llm_refined'] = True
                defn['llm_warnings'] = refined.get('llm_warnings', [])
                llm_refined += 1
                print('OK')
            else:
                defn['llm_refined'] = False
                print('回退规则')

        definitions.append(defn)

    print(f'[GEN] {len(definitions)} 份岗位定义已生成 (LLM精炼: {llm_refined})')

    # 合并同义不同名的定义
    definitions = merge_similar_definitions(definitions)
    print(f'[MERGE] 合并后 {len(definitions)} 份岗位定义')

    return definitions


def dedup_by_jaccard(jds):
    """
    簇内Jaccard抄袭检测（复用 graph_calibrator.py 同样逻辑）。
    同公司+JD文本相似度>0.8 → 抄袭簇 → 每条JD权重=1/簇大小。
    返回每条JD的权重列表。
    """
    n = len(jds)
    texts = [str(r.get('skill_requirements', '')) for r in jds]
    companies = [str(r.get('company_name', '')).strip() for r in jds]
    factors = [1.0] * n   # 簇大小（1=独立JD）

    for i in range(n):
        if factors[i] != 1.0:
            continue  # 已被归入前面的抄袭簇
        cluster = [i]
        for j in range(i + 1, n):
            if factors[j] != 1.0:
                continue
            # 同公司 + 文本相似 > 0.8 = 抄袭
            if companies[i] and companies[i] == companies[j]:
                sim = SequenceMatcher(None, texts[i], texts[j]).ratio()
                if sim > 0.8:
                    cluster.append(j)
        if len(cluster) > 1:
            for idx in cluster:
                factors[idx] = len(cluster)

    # 权重 = 1/簇大小（抄袭簇越大，每条JD贡献越小）
    weights = [1.0 / f for f in factors]
    return weights, factors


def _synthesize_duties_from_skills(all_required, norm_name):
    """从必备技能反推核心职责（JD格式不规范时的兜底）"""
    if not all_required:
        return ['（JD信息不足，无法提取核心职责）']
    # 取最高频技能（保留category信息用于分类）
    skill_info = {}  # {skill: category}
    for sk, _ in all_required:
        name = sk.get('skill', '')
        cat = sk.get('category', '')
        if name not in skill_info:
            skill_info[name] = cat
    top_skills = list(skill_info.keys())[:6]

    ai_skills = [s for s in top_skills if skill_info.get(s) == 'AI新兴技能']
    trad_skills = [s for s in top_skills if skill_info.get(s) != 'AI新兴技能']

    duties = []
    for sk in ai_skills[:3]:
        duties.append(f'负责{sk}相关技术研发与落地，持续跟踪前沿进展并推动业务应用')
    for sk in trad_skills[:3]:
        if sk not in ai_skills[:3]:
            duties.append(f'负责基于{sk}的系统/模块设计、开发与维护')
    if not duties:
        duties.append(f'负责{norm_name}相关技术工作，参与系统架构设计与核心模块开发')
    return duties


def build_definition(cluster_id, norm_name, jds, timeline_index):
    """为一个岗位类型合成定义"""
    n = len(jds)

    # ── 岗位名称 ──
    name_counter = Counter(str(r.get('job_name', '')).strip() for r in jds)
    # 用归一化名作为主名（而非原始名投票—避免把"AI开发工程师"叫成"AI工程师"）
    canonical = norm_name
    # 原始名去重后作为别称（排除与主名完全相同或仅大小写不同的，保留源数据中的多样写法）
    aliases = list(dict.fromkeys(
        nm for nm, _ in name_counter.most_common(15)
        if nm.lower().strip() != canonical.lower().strip()
    ))

    # ── Jaccard抄袭去重（同 graph_calibrator.py） ──
    jd_weights, cluster_factors = dedup_by_jaccard(jds)
    # 统计通胀：簇大小≥3视为重度，2视为中度
    inflate_heavy = sum(1 for f in cluster_factors if f >= 3)
    inflate_medium = sum(1 for f in cluster_factors if f == 2)
    inflate_light = 0  # Jaccard方法只有"抄袭"和"独立"两态
    inflate_clean = sum(1 for f in cluster_factors if f == 1.0)

    # 收集全量文本+技能（Jaccard去重权重）
    all_duties = []
    all_required = []        # (skill_dict, weight)
    all_bonus = []           # (skill_dict, weight)
    all_ai_signals = set()
    companies = []
    cities = []
    duty_counter = Counter()
    total_weight = 0.0       # 有效JD数（去重后）

    for i, r in enumerate(jds):
        w = jd_weights[i]
        total_weight += w

        text = str(r.get('skill_requirements', ''))
        duties = parse_duties(text)
        all_duties.extend(duties)
        for sent in duties:
            for c in range(len(sent) - 1):
                duty_counter[sent[c:c+2]] += w  # 加权

        all_ai_signals |= extract_ai_signals_from_text(text)
        for sk in parse_skill_list(r.get('必备技能', '')):
            all_required.append((sk, w))
        for sk in parse_skill_list(r.get('加分技能', '')):
            # 过滤标注管线的占位文本
            if sk['skill'] not in ('[未识别到明确的加分技能]', '未识别到明确的加分技能', ''):
                all_bonus.append((sk, w))

        company = str(r.get('company_name', '')).strip()
        if company:
            companies.append(company)
        area = str(r.get('work_area', '')).strip()
        if area:
            cities.append(area)

    # ── 核心职责 ──
    core_duties = score_duties(all_duties, all_ai_signals, duty_counter, n)
    if len(core_duties) < 3:
        # 兜底也要去重
        core_duties = []
        seen = set()
        for d in all_duties:
            key = d.strip().lower()
            if key not in seen:
                seen.add(key)
                core_duties.append(d)
            if len(core_duties) >= 5:
                break
    # 兜底：JD格式不规范导致无法提取职责时，从必备技能合成
    if not core_duties:
        core_duties = _synthesize_duties_from_skills(all_required, norm_name)

    # ── 必备技能（Jaccard去重矫正） ──
    req_counter_raw = Counter()    # 原始计数
    req_counter_w = Counter()      # Jaccard去重计数
    req_levels = defaultdict(list)
    for sk, w in all_required:
        req_counter_raw[sk['skill']] += 1
        req_counter_w[sk['skill']] += w
        req_levels[sk['skill']].append(sk['level'])
    required_skills = []
    demoted_to_bonus = []          # 矫正后从必备降级到加分的技能
    effective_n = max(total_weight, 1.0)
    for skill, cnt_w in req_counter_w.most_common():
        raw_support = req_counter_raw[skill] / n
        corrected_support = cnt_w / effective_n
        modal_level = Counter(req_levels[skill]).most_common(1)[0][0]
        cat = 'AI新兴技能' if skill in all_ai_signals else '传统技术'
        skill_entry = {
            'skill': skill, 'category': cat, 'proficiency': modal_level,
            'support_pct': round(raw_support, 2),
            'support_pct_corrected': round(corrected_support, 2),
        }
        if corrected_support >= 0.15 and len(required_skills) < 8:
            required_skills.append(skill_entry)
        elif corrected_support >= 0.06:
            # 矫正后不足以当必备，但够格当加分 → 降级
            demoted_to_bonus.append(skill_entry)
        # else: 矫正后连加分都不够 → 丢弃
        if len(required_skills) >= 8 and len(demoted_to_bonus) >= 6:
            break

    # ── 加分技能（Jaccard去重矫正 + 必备降级） ──
    bonus_counter_raw = Counter()
    bonus_counter_w = Counter()
    bonus_levels = defaultdict(list)
    for sk, w in all_bonus:
        bonus_counter_raw[sk['skill']] += 1
        bonus_counter_w[sk['skill']] += w
        bonus_levels[sk['skill']].append(sk['level'])
    req_skill_names = {s['skill'] for s in required_skills}
    bonus_skills = []

    # 先加入从必备降级下来的技能（already validated）
    for sk_entry in demoted_to_bonus:
        if sk_entry['skill'] not in req_skill_names:
            bonus_skills.append(sk_entry)

    # 再加入原始加分技能（去重必备+已降级）
    existing_bonus = {s['skill'] for s in bonus_skills}
    for skill, cnt_w in bonus_counter_w.most_common():
        if skill in req_skill_names or skill in existing_bonus:
            continue
        raw_support = bonus_counter_raw[skill] / n
        corrected_support = cnt_w / effective_n
        if corrected_support < 0.06:
            break
        modal_level = Counter(bonus_levels[skill]).most_common(1)[0][0]
        cat = 'AI新兴技能' if skill in all_ai_signals else '传统技术'
        bonus_skills.append({
            'skill': skill, 'category': cat, 'proficiency': modal_level,
            'support_pct': round(raw_support, 2),
            'support_pct_corrected': round(corrected_support, 2),
        })
        existing_bonus.add(skill)
        if len(bonus_skills) >= 8:  # 加分上限放宽，容纳降级技能
            break

    # ── 典型行业应用场景（传递JD原文做关键词兜底） ──
    industry_counter = Counter()
    for i, (company, city) in enumerate(zip(companies, cities + [''] * max(0, len(companies) - len(cities)))):
        jd_text = str(jds[i].get('skill_requirements', '')) if i < len(jds) else ''
        industry = infer_industry(company, city, jd_text, jds[i].get('job_name', '') if i < len(jds) else '')
        industry_counter[industry] += 1
    industries = []
    for ind, cnt in industry_counter.most_common(5):
        support = cnt / len(companies) if companies else 0
        if support >= 0.05:
            industries.append({'industry': ind, 'support_pct': round(support, 2)})

    # ── 可溯源 ──
    company_set = list(dict.fromkeys(companies))
    city_set = list(dict.fromkeys(c for c in cities if c))
    # 通胀统计（基于Jaccard抄袭检测，同 graph_calibrator.py）
    inflate_breakdown = {
        '重度(簇≥3)': inflate_heavy,
        '中度(簇=2)': inflate_medium,
        '无': inflate_clean,
    }
    any_inflate = inflate_heavy + inflate_medium
    # 通胀严重度（连续值0~1）：重度JD全权重 + 中度JD半权重
    inflation_severity = round(
        (inflate_heavy * 1.0 + inflate_medium * 0.5) / max(n, 1), 2
    )
    # 通胀比例（连续加权，避免二值化）
    inflation_ratio = inflation_severity  # 直接使用严重度，0=无通胀，1=全部JD严重通胀

    # ── 动态更新 ──
    dynamic = {'linked_timeline': False}
    if timeline_index:
        for job_key, trend in timeline_index.items():
            if normalize_job_name(job_key) == norm_name or canonical in job_key or job_key in canonical:
                dynamic = {
                    'linked_timeline': True,
                    'timeline_report_job_match': job_key,
                    'capability_trend': trend[:200] if trend else '',
                }
                break

    return {
        'cluster_id': cluster_id,
        '岗位名称': canonical,
        '别称': aliases,
        '归一化名': norm_name,
        '核心职责': core_duties,
        '必备技能': required_skills,
        '加分技能': bonus_skills,
        '典型行业应用场景': industries if industries else [{'industry': '通用信息技术', 'support_pct': 1.0}],
        'source_traceability': {
            'total_jds': n,
            'effective_jds': round(total_weight, 1),  # 通胀矫正后有效JD数
            'unique_companies': len(company_set),
            'top_companies': company_set[:5],
            'top_cities': city_set[:5],
            'inflation_ratio': round(inflation_ratio, 2),
            'inflation_severity': inflation_severity,    # 加权严重度(0~1)
            'inflation_breakdown': inflate_breakdown,
        },
        '人工优化': {
            'status': 'pending_review',
            'audit_id': f'newjob_{cluster_id:03d}',
        },
        '动态更新': dynamic,
    }


# ============================================================
#  输出
# ============================================================

import unicodedata

def sanitize_xlsx(val):
    """清洗Excel非法字符（控制字符等）"""
    if not isinstance(val, str):
        return val
    # 移除非法控制字符（保留 \t \n \r）
    cleaned = []
    for ch in val:
        cp = ord(ch)
        if cp < 32 and cp not in (9, 10, 13):
            cleaned.append(' ')
        elif 0xD800 <= cp <= 0xDFFF:
            cleaned.append(' ')  # surrogate pairs
        elif cp == 0xFFFE or cp == 0xFFFF:
            cleaned.append(' ')
        else:
            cleaned.append(ch)
    return ''.join(cleaned)


def export_audit_excel(definitions, output_path):
    """导出待审核定义到Excel，方便人工逐条review"""
    try:
        import openpyxl
    except ImportError:
        print('[WARN] openpyxl未安装，跳过Excel导出')
        return

    wb = openpyxl.Workbook()

    # ── Sheet 1: 待审核定义 ──
    ws1 = wb.active
    ws1.title = '待审核定义'
    headers = [
        '审核编号', '岗位名称', '别称', '归一化名',
        'JD数(原始)', '有效JD(矫正)', '通胀严重度(0~1)', '通胀分级(轻/中/重/无)',
        '核心职责', '必备技能(原始%→矫正%)', '加分技能(原始%→矫正%)',
        '行业场景', '代表公司', '代表城市',
        'LLM精炼', '审核状态', '审核意见'
    ]
    ws1.append(headers)

    # 标题行加粗+底色
    header_fill = openpyxl.styles.PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = openpyxl.styles.Font(color='FFFFFF', bold=True, size=11)
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

    for i, d in enumerate(definitions, 1):
        t = d['source_traceability']
        bd = t.get('inflation_breakdown', {})
        inflate_detail = f"重(簇≥3):{bd.get('重度(簇≥3)',0)} 中(簇=2):{bd.get('中度(簇=2)',0)} 独立:{bd.get('无',0)}"
        row = [
            d['人工优化']['audit_id'],
            d['岗位名称'],
            ' | '.join(d['别称'][:5]),
            d['归一化名'],
            t['total_jds'],
            f"{t['effective_jds']:.0f}",
            f"{t.get('inflation_severity', 0):.2f}",
            inflate_detail,
            '\n'.join(f'{j+1}. {duty}' for j, duty in enumerate(d['核心职责'])),
            '\n'.join(f"{s['skill']} {int(s['support_pct']*100)}%→{int(s.get('support_pct_corrected',s['support_pct'])*100)}%" for s in d['必备技能']),
            '\n'.join(f"{s['skill']} {int(s['support_pct']*100)}%→{int(s.get('support_pct_corrected',s['support_pct'])*100)}%" for s in d['加分技能']),
            '\n'.join(f"{ind['industry']}({int(ind['support_pct']*100)}%)" if 'support_pct' in ind else f"{ind['industry']}({ind.get('rationale','')})" for ind in d['典型行业应用场景']),
            ' | '.join(t['top_companies'][:5]),
            ' | '.join(t['top_cities'][:5]),
            '是' if d.get('llm_refined') else '否',
            '待审核',
            ''  # 审核意见留空
        ]
        row = [sanitize_xlsx(cell) for cell in row]
        ws1.append(row)

    # 列宽
    col_widths = [14, 22, 30, 22, 6, 6, 8, 55, 30, 30, 25, 30, 25, 8, 10, 15]
    for col, w in enumerate(col_widths, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # 冻结首行
    ws1.freeze_panes = 'A2'
    # 自动筛选
    ws1.auto_filter.ref = f'A1:P{len(definitions)+1}'

    # ── Sheet 2: 统计概览 ──
    ws2 = wb.create_sheet('统计概览')
    total_jds = sum(d['source_traceability']['total_jds'] for d in definitions)
    total_companies = sum(d['source_traceability']['unique_companies'] for d in definitions)

    stats = [
        ['指标', '数值'],
        ['定义总数', len(definitions)],
        ['覆盖JD总数', total_jds],
        ['平均JD/定义', round(total_jds / max(len(definitions), 1), 1)],
        ['LLM精炼数', sum(1 for d in definitions if d.get('llm_refined'))],
        ['关联时间轴', sum(1 for d in definitions if d['动态更新']['linked_timeline'])],
        ['', ''],
        ['Top 20 岗位', 'JD数'],
    ]
    for d in sorted(definitions, key=lambda x: -x['source_traceability']['total_jds'])[:20]:
        stats.append([d['岗位名称'], d['source_traceability']['total_jds']])

    for row in stats:
        ws2.append(row)

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 20

    # ── Sheet 3: 行业分布 ──
    ws3 = wb.create_sheet('行业分布')
    industry_all = Counter()
    for d in definitions:
        for ind in d['典型行业应用场景']:
            industry_all[ind['industry']] += d['source_traceability']['total_jds']
    ws3.append(['行业', '关联JD数', '关联定义数'])
    for ind, jd_count in industry_all.most_common():
        def_count = sum(1 for d in definitions
                       if any(i['industry'] == ind for i in d['典型行业应用场景']))
        ws3.append([ind, jd_count, def_count])
    ws3.column_dimensions['A'].width = 25

    wb.save(output_path)
    print(f'[AUDIT] 审核Excel: {output_path}')
    print(f'  待审核定义: {len(definitions)} 条')
    print(f'  3个工作表: 待审核定义 | 统计概览 | 行业分布')



def save_output(definitions, input_dir, output_path, use_llm=False):
    llm_count = sum(1 for d in definitions if d.get('llm_refined'))
    stats = {
        'total_definitions': len(definitions),
        'total_jds_covered': sum(d['source_traceability']['total_jds'] for d in definitions),
        'avg_jds_per_definition': round(sum(d['source_traceability']['total_jds'] for d in definitions) / max(len(definitions), 1), 1),
        'industry_distribution': Counter(i['industry'] for d in definitions for i in d['典型行业应用场景']),
        'definitions_with_timeline': sum(1 for d in definitions if d['动态更新']['linked_timeline']),
        'llm_refined_count': llm_count,
        'engine': 'rule+llm' if use_llm else 'rule',
    }

    output = {
        'generated_at': datetime.now().isoformat(),
        'source': input_dir,
        'num_definitions': len(definitions),
        'definitions': definitions,
        'statistics': stats,
    }

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f'\n[OUTPUT] {output_path}')
    print(f'  岗位定义: {stats["total_definitions"]} 个')
    print(f'  覆盖JD: {stats["total_jds_covered"]} 条')
    print(f'  平均每岗位: {stats["avg_jds_per_definition"]} 条JD')
    print(f'  关联时间轴: {stats["definitions_with_timeline"]} 个')
    print(f'  行业覆盖: {dict(stats["industry_distribution"].most_common(5))}')


# ============================================================
#  CLI
# ============================================================

def main():
    parser = argparse.ArgumentParser(description='新岗位定义生成 — 从标注JD聚合生成结构化岗位定义')
    parser.add_argument('--input_dir', type=str, required=True, help='已标注CSV目录')
    parser.add_argument('--output', type=str, default='new_job_definitions.json', help='输出JSON路径')
    parser.add_argument('--timeline_report', type=str, default='', help='能力动态更新报告（可选）')
    parser.add_argument('--use_llm', action='store_true', help='启用LLM精炼核心职责和行业场景（需DEEPSEEK_API_KEY）')
    parser.add_argument('--llm_max', type=int, default=999, help='LLM精炼最多处理前N个岗位（默认全部）')
    parser.add_argument('--audit_excel', type=str, default='', help='导出待审核Excel路径（如 audit_definitions.xlsx）')
    args = parser.parse_args()

    print('=' * 60)
    engine = '规则+LLM' if args.use_llm else '纯规则引擎'
    print(f'  新岗位定义生成 ({engine})')
    print('=' * 60)

    global pd
    import pandas as pd

    rows = load_new_jobs(args.input_dir)
    if not rows:
        print('[ERROR] 未找到新岗位候选JD')
        sys.exit(1)

    timeline_path = args.timeline_report
    if timeline_path and not os.path.exists(timeline_path):
        # 尝试默认路径
        alt = os.path.join(os.path.dirname(__file__), 'capability_update_report.json')
        if os.path.exists(alt):
            timeline_path = alt

    definitions = group_and_generate(rows, timeline_path,
                                     use_llm=args.use_llm,
                                     llm_max=args.llm_max)
    save_output(definitions, args.input_dir, args.output, use_llm=args.use_llm)

    # 导出审核Excel
    if args.audit_excel:
        export_audit_excel(definitions, args.audit_excel)


if __name__ == '__main__':
    main()
